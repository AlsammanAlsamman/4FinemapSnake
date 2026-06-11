#!/usr/bin/env python3
"""Aggregate per-locus summary TSVs into a single multi-sheet Excel workbook.

Sheet layout
------------
  "Total Summary"  – one row per (target, locus) with best-SNP and count statistics
  "<target>_<locus>" (one per locus) – full per-locus SNP table from summary.tsv

Statistics per locus
--------------------
  COJO   : n_independent_signals, best_snp, best_p, best_or
  FINEMAP: best_snp (highest PIP), best_pip, n_pip_gt_0.25, n_pip_gt_0.5
  SuSiE-R: best_snp (highest PIP), best_pip, n_pip_gt_0.25, n_pip_gt_0.5
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Return a valid Excel sheet name (max 31 chars, no special chars)."""
    for ch in r'\/*?[]':
        name = name.replace(ch, "_")
    return name[:max_len]


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _best_snp_by_max(df: pd.DataFrame, pip_col: str) -> tuple:
    """Return (snp, pip) for the row with the highest pip_col value."""
    valid = df[_numeric(df[pip_col]) > 0].copy()
    valid[pip_col] = _numeric(valid[pip_col])
    if valid.empty:
        return ("", float("nan"))
    idx = valid[pip_col].idxmax()
    return (str(valid.loc[idx, "SNP"]), float(valid.loc[idx, pip_col]))


def _count_above(series: pd.Series, threshold: float) -> int:
    return int((_numeric(series) > threshold).sum())


def _summarize_locus(target: str, locus: str, df: pd.DataFrame) -> dict:
    row: dict = {
        "target": target,
        "locus": locus,
        "chr": df["CHR"].iloc[0] if "CHR" in df.columns and not df.empty else "",
        "locus_start_bp": int(df["BP"].min()) if "BP" in df.columns and not df.empty else "",
        "locus_end_bp": int(df["BP"].max()) if "BP" in df.columns and not df.empty else "",
        "n_snps_tested": len(df),
    }

    # ---- COJO ---------------------------------------------------------------
    cojo_snps = df[df["cojo"].astype(str).str.strip().ne("")].copy() if "cojo" in df.columns else pd.DataFrame()
    row["n_cojo_signals"] = len(cojo_snps)
    if not cojo_snps.empty and "P" in cojo_snps.columns:
        cojo_snps["_p"] = _numeric(cojo_snps["P"])
        best_idx = cojo_snps["_p"].idxmin()
        row["best_cojo_snp"] = str(cojo_snps.loc[best_idx, "SNP"])
        row["best_cojo_p"] = float(cojo_snps.loc[best_idx, "_p"])
        row["best_cojo_or"] = float(cojo_snps.loc[best_idx, "OR"]) if "OR" in cojo_snps.columns else float("nan")
    else:
        row["best_cojo_snp"] = ""
        row["best_cojo_p"] = float("nan")
        row["best_cojo_or"] = float("nan")

    # ---- FINEMAP ------------------------------------------------------------
    if "finemap_pip" in df.columns:
        snp, pip = _best_snp_by_max(df, "finemap_pip")
        row["best_finemap_snp"] = snp
        row["best_finemap_pip"] = pip
        row["n_finemap_pip_gt_0.25"] = _count_above(df["finemap_pip"], 0.25)
        row["n_finemap_pip_gt_0.5"] = _count_above(df["finemap_pip"], 0.50)
    else:
        row.update({"best_finemap_snp": "", "best_finemap_pip": float("nan"),
                    "n_finemap_pip_gt_0.25": 0, "n_finemap_pip_gt_0.5": 0})

    # ---- SuSiE-R ------------------------------------------------------------
    if "susie_pip" in df.columns:
        snp, pip = _best_snp_by_max(df, "susie_pip")
        row["best_susie_snp"] = snp
        row["best_susie_pip"] = pip
        row["n_susie_pip_gt_0.25"] = _count_above(df["susie_pip"], 0.25)
        row["n_susie_pip_gt_0.5"] = _count_above(df["susie_pip"], 0.50)
    else:
        row.update({"best_susie_snp": "", "best_susie_pip": float("nan"),
                    "n_susie_pip_gt_0.25": 0, "n_susie_pip_gt_0.5": 0})

    return row


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")

SECTION_FILLS = {
    "locus_info":  PatternFill("solid", fgColor="2E75B6"),
    "cojo":        PatternFill("solid", fgColor="70AD47"),
    "finemap":     PatternFill("solid", fgColor="ED7D31"),
    "susie":       PatternFill("solid", fgColor="9B59B6"),
}

SECTION_FONTS = {k: Font(bold=True, color="FFFFFF", size=11)
                 for k in SECTION_FILLS}

COLUMN_SECTIONS = {
    "target":               "locus_info",
    "locus":                "locus_info",
    "chr":                  "locus_info",
    "locus_start_bp":       "locus_info",
    "locus_end_bp":         "locus_info",
    "n_snps_tested":        "locus_info",
    "n_cojo_signals":       "cojo",
    "best_cojo_snp":        "cojo",
    "best_cojo_p":          "cojo",
    "best_cojo_or":         "cojo",
    "best_finemap_snp":     "finemap",
    "best_finemap_pip":     "finemap",
    "n_finemap_pip_gt_0.25": "finemap",
    "n_finemap_pip_gt_0.5":  "finemap",
    "best_susie_snp":       "susie",
    "best_susie_pip":       "susie",
    "n_susie_pip_gt_0.25":  "susie",
    "n_susie_pip_gt_0.5":   "susie",
}


def _format_summary_sheet(ws, n_rows: int, n_cols: int) -> None:
    """Apply colour-coded section headers and alternating row fills."""
    cols = [ws.cell(1, c).value for c in range(1, n_cols + 1)]

    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(1, c_idx)
        section = COLUMN_SECTIONS.get(col_name, "locus_info")
        cell.fill = SECTION_FILLS[section]
        cell.font = SECTION_FONTS[section]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, n_rows + 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40


def _format_detail_sheet(ws, n_rows: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, n_rows + 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, n_cols + 1):
            if fill:
                ws.cell(r, c).fill = fill

    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Aggregate per-locus summaries into one Excel workbook")
    parser.add_argument("--results-dir", required=True, help="Top-level results directory")
    parser.add_argument("--out-xlsx", required=True)
    parser.add_argument("--done-file", required=True)
    args = parser.parse_args()

    results_dir = Path(args.results_dir)
    out_xlsx = Path(args.out_xlsx)
    done_file = Path(args.done_file)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    done_file.parent.mkdir(parents=True, exist_ok=True)

    # Discover all per-locus summary TSVs
    summary_files = sorted(results_dir.glob("*/10_summary/*/summary.tsv"))
    if not summary_files:
        print(f"[WARN] No summary.tsv files found under {results_dir}", file=sys.stderr)

    locus_rows = []
    locus_dfs: list[tuple[str, str, pd.DataFrame]] = []

    for tsv_path in summary_files:
        rel = tsv_path.relative_to(results_dir)
        # rel = <target>/10_summary/<locus>/summary.tsv
        target = rel.parts[0]
        locus = rel.parts[2]
        try:
            df = pd.read_csv(tsv_path, sep="\t")
            locus_rows.append(_summarize_locus(target, locus, df))
            locus_dfs.append((target, locus, df))
        except Exception as exc:
            print(f"[WARN] Skipping {tsv_path}: {exc}", file=sys.stderr)

    if not locus_rows:
        # Write an empty workbook rather than failing
        total_df = pd.DataFrame()
    else:
        total_df = pd.DataFrame(locus_rows)

    # Write Excel workbook
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # Sheet 1: Total Summary
        total_df.to_excel(writer, sheet_name="Total Summary", index=False)
        ws = writer.sheets["Total Summary"]
        _format_summary_sheet(ws, len(total_df), len(total_df.columns))

        # Per-locus sheets (one per locus)
        used_sheet_names: dict[str, int] = {}
        for target, locus, df in locus_dfs:
            raw_name = f"{target}_{locus}"
            base = _safe_sheet_name(raw_name)
            # Deduplicate sheet names
            if base in used_sheet_names:
                used_sheet_names[base] += 1
                sheet_name = _safe_sheet_name(f"{base}_{used_sheet_names[base]}", 31)
            else:
                used_sheet_names[base] = 0
                sheet_name = base

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_locus = writer.sheets[sheet_name]
            _format_detail_sheet(ws_locus, len(df), len(df.columns))

    with open(done_file, "w", encoding="utf-8") as handle:
        handle.write("ok\n")

    print(f"Written {len(locus_rows)} loci → {out_xlsx}")


if __name__ == "__main__":
    main()
